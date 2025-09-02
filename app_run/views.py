from django.contrib.auth.models import User
from django.db.models import Sum
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from geopy.distance import geodesic
from openpyxl.reader.excel import load_workbook
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from django.conf import settings
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView

from app_run.models import Run, AthleteInfo, Challenge, Position, CollectibleItem
from app_run.serializers import RunSerializer, UserSerializer, AthleteInfoSerializer, ChallengeSerializer, \
    PositionSerializer, CollectibleItemSerializer


@api_view(['GET'])
def view_about(request):
    details = {
        'company_name': settings.COMPANY_NAME,
        'slogan': settings.SLOGAN,
        'contacts': settings.CONTACTS
    }
    return Response(details)


class RunPagination(PageNumberPagination):
    page_size_query_param = 'size'
    max_page_size = 50


class UserPagination(PageNumberPagination):
    page_size_query_param = 'size'
    max_page_size = 50


class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.all().select_related('athlete')
    serializer_class = RunSerializer
    pagination_class = RunPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    ordering = ['id']  # Сортировка по умолчанию


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.filter(is_superuser=False)
    serializer_class = UserSerializer
    pagination_class = UserPagination
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']

    def get_queryset(self):
        qs = self.queryset
        type = self.request.query_params.get('type')
        if type == 'coach':
            qs = qs.filter(is_staff=True)
        if type == 'athlete':
            qs = qs.filter(is_staff=False)
        return qs


class RunStartAPIView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)

        if run.status != 'init':
            return Response({'detail': 'Invalid run status for starting.'}, status=status.HTTP_400_BAD_REQUEST)

        run.status = 'in_progress'
        run.save()

        return Response(RunSerializer(run).data, status=status.HTTP_201_CREATED)


class RunStopAPIView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)

        if run.status != 'in_progress':
            return Response({'detail': 'Invalid run status for starting.'}, status=status.HTTP_400_BAD_REQUEST)

        user = run.athlete

        positions = list(run.positions.all().order_by('id'))

        if len(positions) < 2:
            run.status = 'finished'
            run.save()
            return Response({'error': 'Run stopped.  Not enough positions to calculate distance.'},
                            status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        total_distance = 0
        previous_coordinate = None

        for position in positions:
            coordinates = (position.latitude, position.longitude)

            if previous_coordinate is not None:
                distance = geodesic(previous_coordinate, coordinates).kilometers
                total_distance += distance

            previous_coordinate = coordinates

        run.distance = round(total_distance, 3)
        run.status = 'finished'
        run.save(update_fields=['status', 'distance'])

        sum_distance = Run.objects.filter(athlete=user, status='finished').aggregate(total_sum=Sum('distance'))[
                           'total_sum'] or 0

        try:
            athlete_info = user.athleteinfo
        except AthleteInfo.DoesNotExist:
            return Response({'detail': 'AthleteInfo not found'}, status=status.HTTP_400_BAD_REQUEST)

        if Run.objects.filter(athlete=user, status='finished').count() == 10:
            if not Challenge.objects.filter(athlete=athlete_info, full_name='Сделай 10 Забегов!').exists():
                Challenge.objects.create(athlete=athlete_info, full_name='Сделай 10 Забегов!')

        if sum_distance >= 50:
            if not Challenge.objects.filter(athlete=athlete_info, full_name='Пробеги 50 километров!').exists():
                Challenge.objects.create(athlete=athlete_info, full_name='Пробеги 50 километров!')

        return Response(RunSerializer(run).data, status=status.HTTP_200_OK)


class AthleteInfoView(APIView):
    def get(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        athlete, created = AthleteInfo.objects.get_or_create(user=user)
        serializer = AthleteInfoSerializer(athlete)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        weight = request.data.get('weight')
        goals = request.data.get('goals')

        if weight is None or goals is None:
            return Response({'error': 'Missing data!'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            weight = int(weight)
        except ValueError:
            return Response({'error': 'Invalid weight format!'}, status=status.HTTP_400_BAD_REQUEST)

        if weight <= 0 or weight >= 900:
            return Response({'error': 'Invalid weight!'}, status=status.HTTP_400_BAD_REQUEST)

        athlete, created = AthleteInfo.objects.update_or_create(user=user,
                                                                defaults={
                                                                    'weight': weight,
                                                                    'goals': goals
                                                                })
        serializer = AthleteInfoSerializer(athlete)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ChallengeViewSet(viewsets.ModelViewSet):
    queryset = Challenge.objects.all()
    serializer_class = ChallengeSerializer

    def get_queryset(self):
        qs = self.queryset
        athlete_id = self.request.query_params.get('athlete')
        if athlete_id:
            qs = qs.filter(athlete__id=athlete_id)
            return qs
        return qs


class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = Position.objects.all()
        run_id = self.request.query_params.get('run', None)
        if run_id:
            qs = qs.filter(run=run_id)
            return qs
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        position_id = serializer.data['id']
        headers = self.get_success_headers(serializer.data)
        return Response({'id': position_id}, status=status.HTTP_201_CREATED, headers=headers)


class CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer

class UploadFileView(APIView):
    def post(self, request):
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            return Response({'error': 'File not found.'}, status=status.HTTP_400_BAD_REQUEST)

        if not uploaded_file.name.endswith('.xlsx'):
            return Response({'error': 'File is not xlsx.'}, status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)

        wb = load_workbook(filename=uploaded_file, data_only=True)
        #wb = load_workbook(r'C:\Users\Pavel\Downloads\upload_example.xlsx')
        ws = wb.active
        invalid_rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_list = list(row)

            data = {
                'name': row[0],
                'uid': row[1],
                'value': row[2],
                'latitude': row[3],
                'longitude': row[4],
                'picture': row[5],
            }
            if CollectibleItem.objects.filter(uid=data['uid']).exists():
                continue

            serializer = CollectibleItemSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                invalid_rows.append(row_list)

        return Response(invalid_rows, status=status.HTTP_201_CREATED)
